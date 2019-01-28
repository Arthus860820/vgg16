import tkinter as tk
import time
import os
import openpyxl as op
from openpyxl import Workbook


# 定義觸發事件時的函式create_new_totalfile（注意：因為Python的執行順序是從上往下，所以函式一定要放在按鈕的上面
def create_new_totalfile(work_sheet,value):
    var=value+'.xlsx'
    if os.path.isfile(var):
        window_warning=tk.Tk()
        window_warning.title('warning')
        window_warning.geometry('400x80')
        n=tk.Label(window_warning,fg='red',text=var+'已經存在，你是否要覆蓋它',font=('DFKai-SB', 14))
        n.place(x=0,y=0,width=400,height=30)
        def yes():
            #刪除window_warning視窗
            window_warning.destroy()
            workbooknew=Workbook()
            sheet2= workbooknew.active
            sheet2.title=value
            w=tk.Label(window,bg='green',fg='white',font=('DFKai-SB', 14),text='successful:以建立一個新檔案:'+var)
            w.place(x=0,y=30,width=1000,height=30)
            max_row=work_sheet.max_row#最大行數
            max_column=work_sheet.max_column#最大列數 
            for m in range(1,max_row+1):
                for n in range(97,97+max_column):#chr(97)='a'
                    n=chr(n)#ASCII字元
                    i='%s%d'%(n,m)#單元格編號
                    cell1=work_sheet[i].value#獲取data單元格資料
                    sheet2[i].value=cell1#賦值到test單元格
            #存檔
            workbooknew.save(var)
            #建立按鈕
        y=tk.Button(window_warning,fg='red',text='清除並建立',font=('DFKai-SB', 16),command=yes)#command=函式名
        y.place(x=0,y=30,width=180,height=30)
        def no():
            #刪除window_warning視窗
            window_warning.destroy()
        y=tk.Button(window_warning,fg='green',text='取消/重新取名',font=('DFKai-SB', 16),command=no)#command=函式名
        y.place(x=200,y=30,width=180,height=30)
        window_warning.mainloop()
        
    else:
        workbooknew=Workbook()
        sheet2= workbooknew.active
        sheet2.title=value
        w=tk.Label(window,bg='green',fg='white',font=('DFKai-SB', 14),text='successful:以建立一個新檔案:'+var)
        w.place(x=0,y=30,width=1000,height=30)
        max_row=work_sheet.max_row#最大行數
        max_column=work_sheet.max_column#最大列數 
        for m in range(1,max_row+1):
            for n in range(97,97+max_column):#chr(97)='a'
                n=chr(n)#ASCII字元
                i='%s%d'%(n,m)#單元格編號
                cell1=work_sheet[i].value#獲取data單元格資料
                sheet2[i].value=cell1#賦值到test單元格
        #存檔
        workbooknew.save(var)
        

#建立視窗
window=tk.Tk() 
#視窗名子
window.title('name_to_xls') 


#視窗大小中間的*是小x
window.geometry('1000x800') 

#在GUI上設定一個標籤用來放字
n=tk.Label(window,font=('DFKai-SB', 16),text='檔案名稱(含副檔名):')
# 放置lable的方法有：1）l.pack()2.place()
n.place(x=5,y=0,width=205,height=30)
# 在圖形介面上設定輸入框控制元件entry框並放置

e=tk.Entry(window,show=None, font=('Arial', 16))#顯示成明文形式
e.place(x=215,y=0,width=245,height=30)

#先建立變數在函式外，讓函式執行完後這個變數跟他的值還存在
var_allsheet= tk.StringVar()
mylist_allsheet= tk.StringVar()

def read_old_totalfile():
    #時間
    #取得e的框框中的文字
    var=e.get()
    #檢查檔案是否存在
    if os.path.isfile(var):    
        w=tk.Label(window,bg='green',fg='white',font=('DFKai-SB', 14),text='successful:已成功讀取檔案'+var)
        w.place(x=0,y=30,width=1000,height=30)
        #讀取excel
        workbook=op.load_workbook(var)
        #將所有的sheet存到allsheet
        allsheet=workbook.sheetnames
        print(allsheet)
        var_allsheet.set(allsheet)
        #建立一個sheet的list table
        lb=tk.Listbox(window,listvariable=var_allsheet)
        lb.place(x=5,y=60,width=200,height=300)
        
        def copy_worksheet():
            value = lb.get(lb.curselection()) #讀滑鼠選到的值
            create_new_totalfile(workbook[value],value)
                
        c=tk.Button(window,text='複製並建立的工作表',font=('DFKai-SB',14),command=copy_worksheet)
        c.place(x=210,y=60,width=200,height=30)
        #建立一個data的 list table
        def seedata():
            value = lb.get(lb.curselection())
            workbook_sheet=workbook[value]
            #建立一個陣列放等等的輸出
            mylist = []
            count=0#用來計算第幾個人(因為下面的row的格式不是數字qq)
            for row in workbook_sheet.iter_rows('A{}:B{}'.format(workbook_sheet.min_row,workbook_sheet.max_row)):
                count+=1
                data=str(count)+' '
                for cell in row:
                    data=data+str(cell.value)+'  '
                mylist.append(data)
            #將資料換格式
            mylist_allsheet.set(mylist)
            ld=tk.Listbox(window,listvariable=mylist_allsheet)
            ld.place(x=500,y=60,width=200,height=300)
            maxrow=workbook_sheet.max_row
            num_last=maxrow
            name_last=workbook_sheet['A'+str(maxrow)].value
            time_last=workbook_sheet['B'+str(maxrow)].value
            w=tk.Label(window,bg='blue',fg='white',font=('DFKai-SB', 14),text='最後一筆資料:第'+str(num_last)+'位 名子:'+str(name_last)+'   登入時間'+str(time_last))
            w.place(x=0,y=30,width=1000,height=30)
            
        c=tk.Button(window,text='查看該工作表內資料',font=('DFKai-SB',14),command=seedata)
        c.place(x=210,y=90,width=200,height=30)
        
        def add_data():
            value = lb.get(lb.curselection())
            workbook_sheet=workbook[value]
            maxrow=workbook_sheet.max_row
            window_add_data=tk.Tk()
            window_add_data.title('新增資料')
            window_add_data.geometry('400x120')
            #時間
            time_now=time.strftime("%Y-%m-%d %H:%M:%S")
            n=tk.Label(window_add_data,fg='black',text='第'+str(maxrow+1)+'位輸入者',font=('DFKai-SB', 14))
            n.place(x=0,y=0,width=400,height=30)
            n=tk.Label(window_add_data,fg='black',text='名子:',font=('DFKai-SB', 14))
            n.place(x=0,y=30,width=70,height=30)
            en=tk.Entry(window_add_data,show=None, font=('DFKai-SB', 14))#顯示成明文形式
            en.place(x=75,y=30,width=325,height=30)
            n=tk.Label(window_add_data,fg='black',text='時間'+time_now,font=('DFKai-SB', 14))
            n.place(x=0,y=60,width=400,height=30)
            
            def add_yes():
                envalue=en.get()
                #如果 value.xlsx存在 新增資料進去
                if os.path.isfile(value+'.xlsx'):
                    wb=op.load_workbook(value+'.xlsx')
                    ws=wb[value]
                    maxRow=ws.max_row
                    ws['A'+str(maxRow+1)]=str(envalue)
                    print(maxRow)
                    ws['B'+str(maxRow+1)]=str(time_now)
                    wb.save(value+'.xlsx')
                #新增資料到record.xlsx的value worksheet中
                workbook_sheet['A'+str(maxrow+1)]=str(envalue)
                print(maxrow,envalue)
                workbook_sheet['B'+str(maxrow+1)]=str(time_now)
                window_add_data.destroy()
                seedata()
                workbook.save(var)
                
            n=tk.Button(window_add_data,text='確定新增',font=('DFKai-SB',14),command=add_yes)
            n.place(x=0,y=90,width=200,height=30)
            
            def cancel():
                window_add_data.destroy()
                
            n=tk.Button(window_add_data,text='取消',font=('DFKai-SB',14),command=cancel)
            n.place(x=200,y=90,width=200,height=30)
            window_add_data.mainloop()
            
        c=tk.Button(window,text='新增資料到尾端',font=('DFKai-SB',14),command=add_data)
        c.place(x=710,y=60,width=200,height=30)
    else:
        #在GUI上設定一個標籤用來放字
        w=tk.Label(window,bg='red',fg='white',font=('DFKai-SB', 14),text='warning:找不到檔案 請確定檔案是否與此程式檔在同一個資料夾或都在桌面')
        w.place(x=0,y=30,width=1000,height=30)

b=tk.Button(window,text='讀取舊檔',font=('DFKai-SB',14),command=read_old_totalfile)
b.place(x=465,y=0,width=195,height=30)


# 注意，loop因為是迴圈的意思，window.mainloop就會讓window不斷的重新整理，如果沒有mainloop,就是一個靜態的window,傳入進去的值就不會有迴圈，mainloop就相當於一個很大的while迴圈，有個while，每點選一次就會更新一次，所以我們必須要有迴圈
window.mainloop()
